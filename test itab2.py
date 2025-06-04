import configparser
import os
import re
import xlwings as wx
import win32com.client
from openpyxl.reader.excel import load_workbook

from secondary_func import get_path_itab_file, create_export_for_itab

project_path = r"C:\Users\Denis.Reshilov\Desktop\app_tiburon\test\itabtest1\Programming\data\live"
project_number = "2306243301"
itab_macro_name = "testLoadDic"
sheet_name_dict = "dict"
itab_path = r"C:\Users\Denis.Reshilov\Desktop\app_tiburon\test\itabtest1\Task\itab_v28_23-000000-01.xlsm"

export_wb = load_workbook(f"{project_path}\\export.xlsx")
export_ws = export_wb["Sheet1"]
for index, cell in enumerate(export_ws['A']):
    if cell.value == "Variable Information":
        export_ws.delete_rows(1, index)
        export_wb.save(f"{project_path}\\export.xlsx")

itab_file_wb = load_workbook(itab_path, keep_vba=True)
itab_file_ws = itab_file_wb[sheet_name_dict]
itab_file_ws.delete_cols(0, 100)

for index_row, row in enumerate(export_ws.iter_rows()):
    for index_column, cell in enumerate(row):
        itab_file_ws.cell(row=index_row + 1, column=index_column + 1, value=cell.value)

export_wb.close()
itab_file_wb.save(itab_path)
itab_file_wb.close()

with wx.App(visible=False) as app:
    vba = app.books.open(itab_path)
    vba_macro = vba.macro(itab_macro_name)
    vba_macro()
    vba.save(itab_path)
