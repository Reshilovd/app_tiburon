import os
import subprocess
from openpyxl import load_workbook
import zipfile
import xlwings as wx
import configparser

def get_path_itab_file(project_path):
    if os.path.exists("new_itab_path.txt"):
        with open("new_itab_path.txt") as file:
            new_itab_path = file.read()
            return new_itab_path

    if os.path.exists("settings.ini"):
        config = configparser.ConfigParser()
        config.read("settings.ini")
        itab_name = config["Template"]["template_name"]

        itab_path = project_path.split("Programming")[0]
        if "Task" in os.listdir(itab_path):
            itab_path = itab_path + "Task"
            if itab_name in os.listdir(itab_path):
                return itab_path + f'\\{itab_name}'


def get_itab(project_path, project_number, itab_macro_name, sheet_name_dict):
    print(3)
    itab_path = get_path_itab_file(project_path)
    print(4)
    if itab_path is None:
        return 0

    if len(itab_path.split("\\")) == 0:
        itab_name = itab_path.split("/")[-1]
    else:
        itab_name = itab_path.split("\\")[-1]

    itab_version = itab_name.split("itab_v")[1].split('_')[0]

    try:
        itab_file_wb = load_workbook(itab_path, keep_vba=True)
        itab_file_ws = itab_file_wb.active
    except FileNotFoundError:
        return 0

    try:
        dict_sheet = itab_file_wb[sheet_name_dict]
        dict_sheet.delete_cols(0, 100)
    except KeyError:
        return 1

    syntax_sps_file = open(f"{project_path}\\itab_export.sps", "w")
    syntax_sps = f"""new file.
    set unicode = on.
    get file = "{project_path}\\{project_number}.sav".
    DISPLAY DICTIONARY.
    OUTPUT EXPORT
      /CONTENTS  EXPORT=VISIBLE  LAYERS=PRINTSETTING  MODELVIEWS=PRINTSETTING
      /XLSM  DOCUMENTFILE='{project_path}\\export.xlsm'
         OPERATION=CREATEFILE
         LOCATION=LASTCOLUMN  NOTESCAPTIONS=YES."""

    syntax_sps_file.write(syntax_sps)
    syntax_sps_file.close()
    print(5)

    itab_spj_file = open(f"{project_path}\\itab_task.spj", "w")
    itab_syntax_spj = (f'<?xml version="1.0" encoding="UTF-8"?><job codepageSyntaxFiles="false" print="true" '
                       f'syntaxErrorHandling="continue" syntaxFormat="interactive" unicode="true" '
                       f'xmlns="http://www.ibm.com/software/analytics/spss/xml/production" '
                       f'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                       f'xsi:schemaLocation="http://www.ibm.com/software/analytics/spss/xml/production '
                       f'http://www.ibm.com/software/analytics/spss/xml/production/production-1.4.xsd"><locale '
                       f'charset="UTF-8" country="RU" language="ru"/><output outputFormat="viewer" '
                       f'outputPath="test.spv"/><syntax syntaxPath="{project_path}\\itab_export.sps"/></job>')
    itab_spj_file.write(itab_syntax_spj)
    itab_spj_file.close()
    print(6)

    itab_run_bat = open(f"{project_path}\\itab_run.bat", "w")
    itab_run_bat.write(
        f'"C:\\Program Files\\IBM\\SPSS\\Statistics\\24\\stats.exe" "{project_path}\\itab_task.spj" -production silent')
    itab_run_bat.close()
    print(7)

    subprocess.run(f'{project_path}\\itab_run.bat', shell=True)
    print(8)
    try:
        itab_wb = load_workbook(f"{project_path}\\export.xlsm", keep_vba=True)
        itab_ws = itab_wb.active
        for index, cell in enumerate(itab_ws['A']):
            if cell.value == "Variable Information":
                itab_ws.delete_rows(1, index)
                try:
                    itab_wb.save(f"{project_path}\\new_export.xlsm")
                    itab_wb.close()
                    break
                except PermissionError:
                    os.remove(f"{project_path}\\itab_run.bat")
                    os.remove(f"{project_path}\\itab_task.spj")
                    os.remove(f"{project_path}\\export.xlsm")
                    os.remove(f"{project_path}\\new_export.xlsm")
                    os.remove(f"{project_path}\\itab_export.sps")
                    return 2

        for index_row, row in enumerate(itab_ws.iter_rows()):
            for index_column, cell in enumerate(row):
                dict_sheet.cell(row=index_row + 1, column=index_column + 1, value=cell.value)
    except FileNotFoundError as ex:
        print(ex)

    print(9)

    try:
        itab_file_wb.save(itab_path)
        itab_file_wb.close()
    except PermissionError:
        os.remove(f"{project_path}\\itab_run.bat")
        os.remove(f"{project_path}\\itab_task.spj")
        os.remove(f"{project_path}\\export.xlsm")
        os.remove(f"{project_path}\\new_export.xlsm")
        os.remove(f"{project_path}\\itab_export.sps")
        return 2

    if os.path.exists(itab_path):
        vba = wx.Book(itab_path)
        try:
            vba_macro = vba.macro(itab_macro_name)
            vba_macro()
            vba.save()
            vba.close()
        except Exception:
            return 4
    else:
        print(f"{itab_path} не найден!")

    if len("\\".join(itab_path.split("\\")[0:-1])) == 0:
        new_itab_path = "/".join(itab_path.split("/")[0:-1]) + f"/itab_v{itab_version}_{project_number}.xlsm"
    else:
        new_itab_path = "\\".join(itab_path.split("\\")[0:-1]) + f"/itab_v{itab_version}_{project_number}.xlsm"
    print(new_itab_path)
    try:
        os.rename(itab_path, new_itab_path)
    except FileExistsError:
        return 3
    finally:
        os.remove(f"{project_path}\\itab_run.bat")
        os.remove(f"{project_path}\\itab_task.spj")
        os.remove(f"{project_path}\\export.xlsm")
        os.remove(f"{project_path}\\new_export.xlsm")
        os.remove(f"{project_path}\\itab_export.sps")

    return new_itab_path


out = get_itab(r"C:\Users\Denis.Reshilov\Desktop\app_tiburon\test\itabtest2\Programming\data\live",
               "2306243301",
               itab_macro_name="testLoadDic",
               sheet_name_dict="dict")
print(out)