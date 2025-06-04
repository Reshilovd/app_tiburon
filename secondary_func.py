import configparser
import glob
import os
import subprocess
import zipfile
import time
import traceback
import xlwings as wx
import pandas as pd
import math
import re
import win32com.client
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def error_checker(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception:
            with open("logs/logs.txt", "a+") as file:
                traceback_info = traceback.format_exc()
                file.write("\n" + datetime.now().strftime('%Y-%m-%d %H:%M'))
                file.write("\n" + traceback_info)
                file.write("\n" + "------------------------------")

    return wrapper


def get_shortcut_target_path(shortcut_path):
    shell = win32com.client.Dispatch("WScript.Shell")
    shortcut = shell.CreateShortCut(shortcut_path)
    target_path = shortcut.TargetPath
    return target_path


def create_export_for_itab(project_path, project_number):
    syntax_sps_file = open(f"{project_path}\\itab_export.sps", "w")
    syntax_sps = f"""new file.
        set unicode = on.
        get file = "{project_path}\\{project_number}.sav".
        DISPLAY DICTIONARY.
        OUTPUT EXPORT
          /CONTENTS  EXPORT=VISIBLE  LAYERS=PRINTSETTING  MODELVIEWS=PRINTSETTING
          /XLSX  DOCUMENTFILE='{project_path}\\export.xlsx'
             OPERATION=CREATEFILE
             LOCATION=LASTCOLUMN  NOTESCAPTIONS=YES."""

    syntax_sps_file.write(syntax_sps)
    syntax_sps_file.close()

    itab_spj_file = open(f"{project_path}\\itab_task.spj", "w")
    itab_syntax_spj = (f'<?xml version="1.0" encoding="UTF-8"?><job codepageSyntaxFiles="false" print="true" '
                       f'syntaxErrorHandling="continue" syntaxFormat="interactive" unicode="true" '
                       f'xmlns="http://www.ibm.com/software/analytics/spss/xml/production" '
                       f'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                       f'xsi:schemaLocation="http://www.ibm.com/software/analytics/spss/xml/production '
                       f'http://www.ibm.com/software/analytics/spss/xml/production/production-1.4.xsd"><locale '
                       f'charset="UTF-8" country="RU" language="ru"/><output outputFormat="viewer" '
                       f'outputPath="test.spv"/><syntax syntaxPath="{project_path}\\itab_export.sps"/></job>')
    itab_spj_file.write(itab_syntax_spj)
    itab_spj_file.close()

    itab_run_bat = open(f"{project_path}\\itab_run.bat", "w")
    itab_run_bat.write(
        f'"C:\\Program Files\\IBM\\SPSS\\Statistics\\24\\stats.exe" "{project_path}\\itab_task.spj" -production '
        f'silent')
    itab_run_bat.close()

    subprocess.run(f'{project_path}\\itab_run.bat', shell=True)

    os.remove(f"{project_path}\\itab_run.bat")
    os.remove(f"{project_path}\\itab_task.spj")
    os.remove(f"{project_path}\\itab_export.sps")


@error_checker
def formatting_path(path):
    valid_path = path.replace("\\", "\\\\")
    if valid_path[0] == '"' or valid_path[0] == "'":
        valid_path = valid_path[1:]
    if valid_path[-1] == '"' or valid_path[-1] == "'":
        valid_path = valid_path[0:-1]
    return valid_path


@error_checker
def conversion_base_in_sav(project_number, project_path):
    syntax_sps = open(f"{project_path}\\{project_number}.sps", "r").read()
    syntax_sps = syntax_sps.replace('/NAME="".', f'/NAME="{project_path}".')

    new_syntax_sps = open(f"{project_path}\\{project_number}.sps", "w")
    new_syntax_sps.write(syntax_sps)
    new_syntax_sps.close()

    spj_file = open(f"{project_path}\\task.spj", "w")
    syntax_spj = (f'<?xml version="1.0" encoding="UTF-8"?><job codepageSyntaxFiles="false" print="true" '
                  f'syntaxErrorHandling="continue" syntaxFormat="interactive" unicode="true" '
                  f'xmlns="http://www.ibm.com/software/analytics/spss/xml/production" '
                  f'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                  f'xsi:schemaLocation="http://www.ibm.com/software/analytics/spss/xml/production '
                  f'http://www.ibm.com/software/analytics/spss/xml/production/production-1.4.xsd"><locale '
                  f'charset="UTF-8" country="RU" language="ru"/><output outputFormat="viewer" '
                  f'outputPath="test.spv"/><syntax syntaxPath="{project_path}\\{project_number}.sps"/></job>')
    spj_file.write(syntax_spj)
    spj_file.close()

    run_bat = open(f"{project_path}\\run.bat", "w")
    run_bat.write(
        f'"C:\\Program Files\\IBM\\SPSS\\Statistics\\24\\stats.exe" "{project_path}\\task.spj" -production silent')
    run_bat.close()

    subprocess.run(f'{project_path}\\run.bat', shell=True)

    os.remove(f"{project_path}\\run.bat")
    os.remove(f"{project_path}\\task.spj")

    return 1


@error_checker
def move_and_unzip(project_path, date_now, user_name):
    list_of_files = glob.glob(
        f"C:\\Users\\{user_name}\\Downloads\\*.zip")
    latest_file = max(list_of_files, key=os.path.getctime)

    stat = os.stat(latest_file)
    ctime = stat.st_ctime
    date_create_base = datetime.fromtimestamp(ctime)

    if latest_file is not None:
        if date_now < date_create_base:
            with zipfile.ZipFile(latest_file, 'r') as base_zip:
                base_zip.extractall(project_path)

                for file in base_zip.namelist():
                    if ".csv" in file:
                        project_number = file.replace(".csv", "")
                        return project_number

        else:
            time.sleep(10)
            return move_and_unzip(project_path, date_now, user_name)


@error_checker
def get_path_itab_file(project_path, itab_name):
    itab_path = project_path.split("Programming")[0]
    if "Task" in os.listdir(itab_path):
        itab_path = itab_path + "Task"
        if itab_name in os.listdir(itab_path):
            return itab_path + f'\\{itab_name}'


def get_itab(project_path, project_number, itab_macro_name, sheet_name_dict):
    try:
        config = configparser.ConfigParser()
        config.read("settings.ini")
        itab_name = config["Template"]["template_name"]
        if not os.path.exists(f"{project_path}\\{project_number}.sav"):
            return 5

        itab_path = get_path_itab_file(project_path, itab_name)
        if itab_path is None:
            return 0

        new_itab_path = re.sub(r'..-000000-..', project_number, itab_path)

        if os.path.exists(new_itab_path):
            return 3

        create_export_for_itab(project_path, project_number)

        export_wb = load_workbook(f"{project_path}\\export.xlsx")
        export_ws = export_wb["Sheet1"]
        for index, cell in enumerate(export_ws['A']):
            if cell.value == "Variable Information":
                export_ws.delete_rows(1, index)
                export_wb.save(f"{project_path}\\export.xlsx")

        itab_file_wb = load_workbook(itab_path, keep_vba=True)
        itab_file_ws = itab_file_wb[sheet_name_dict]
        itab_file_ws.delete_cols(0, 100)

        for index_row, row in enumerate(export_ws.iter_rows()):
            for index_column, cell in enumerate(row):
                itab_file_ws.cell(row=index_row + 1, column=index_column + 1, value=cell.value)

        export_wb.close()
        itab_file_wb.save(itab_path)
        itab_file_wb.close()

        with wx.App(visible=False) as app:
            vba = app.books.open(itab_path)
            vba_macro = vba.macro(itab_macro_name)
            vba_macro()
            vba.save(itab_path)

        # vba = wx.Book(itab_path)
        # vba_macro = vba.macro(itab_macro_name)
        # vba_macro()
        # vba.save(itab_path)
        # vba.close()

        # excel_app = win32com.client.Dispatch("Excel.Application")
        # wb = excel_app.Workbooks.Open(itab_path)
        # excel_app.Application.Run(itab_macro_name)
        # wb.Save()
        # wb.Close()

        os.rename(itab_path, new_itab_path)

        return new_itab_path
    except FileNotFoundError as ex:
        with open("logs/logs.txt", "a+") as file:
            traceback_info = traceback.format_exc()
            file.write("\n" + datetime.now().strftime('%Y-%m-%d %H:%M'))
            file.write("\n" + traceback_info)
            file.write("\n" + str(ex))
            file.write("\n" + "------------------------------")
        return 0
    except KeyError as ex:
        with open("logs/logs.txt", "a+") as file:
            traceback_info = traceback.format_exc()
            file.write("\n" + datetime.now().strftime('%Y-%m-%d %H:%M'))
            file.write("\n" + traceback_info)
            file.write("\n" + str(ex))
            file.write("\n" + "------------------------------")
        return 1
    except PermissionError as ex:
        with open("logs/logs.txt", "a+") as file:
            traceback_info = traceback.format_exc()
            file.write("\n" + datetime.now().strftime('%Y-%m-%d %H:%M'))
            file.write("\n" + traceback_info)
            file.write("\n" + str(ex))
            file.write("\n" + "------------------------------")
        return 2
    except FileExistsError as ex:
        with open("logs/logs.txt", "a+") as file:
            traceback_info = traceback.format_exc()
            file.write("\n" + datetime.now().strftime('%Y-%m-%d %H:%M'))
            file.write("\n" + traceback_info)
            file.write("\n" + str(ex))
            file.write("\n" + "------------------------------")
        return 3
    # это ловлю ошибку, которая возникает, когда не подключен диск М, на котором лежит файл для макроса шаблона
    except Exception as ex:
        with open("logs/logs.txt", "a+") as file:
            traceback_info = traceback.format_exc()
            file.write("\n" + datetime.now().strftime('%Y-%m-%d %H:%M'))
            file.write("\n" + traceback_info)
            file.write("\n" + str(ex))
            file.write("\n" + "------------------------------")
        return 4
    finally:
        if os.path.exists(f"{project_path}\\export.xlsx"):
            os.remove(f"{project_path}\\export.xlsx")


@error_checker
def get_final_id(project_path, project_number, panel):
    if not os.path.exists(f"{project_path}\\{project_number}.sav"):
        return 1
    df = pd.read_spss(f"{project_path}\\{project_number}.sav")
    new_wb = Workbook()
    panel_list = []
    # and len(df[panel].dropna())
    # df = df.fillna({'panel': "-"})
    if panel in df.columns:
        for value in df[panel]:
            if isinstance(value, float):
                if math.isnan(value):
                    if "None" not in panel_list:
                        panel_list.append("None")
            else:
                if value not in panel_list:
                    panel_list.append(value)

        for index, panel_name in enumerate(panel_list):
            if not isinstance(panel_name, str):
                panel_list.pop(index)

        for panel_name in panel_list:
            if panel_name == "None":
                new_df = df[df[panel].isnull()]
            else:
                new_df = df.loc[df[panel] == panel_name]

            new_wb.create_sheet(panel_name)
            sheet = new_wb[panel_name]

            number_rows = new_df.shape[0]

            sheet.cell(row=1, column=1, value="№")
            sheet.cell(row=1, column=2, value="InterviewUID")
            sheet.cell(row=1, column=3, value="Status")
            sheet.cell(row=1, column=4, value=panel)
            sheet.cell(row=1, column=5, value="Total:")
            sheet.cell(row=1, column=6, value=number_rows)

            for index, row in enumerate(new_df.itertuples()):

                # if row.panel.isnull():

                sheet.append([index + 1, row.InterviewUID, row.Status, row.panel])

                # подгонка ширины столбцов
                for column in sheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    adjusted_width = (max_length + 2)  # добавляем небольшой запас
                    sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    else:

        new_wb.create_sheet("panel")
        sheet = new_wb["panel"]
        number_rows = df.shape[0]

        sheet.cell(row=1, column=1, value="№")
        sheet.cell(row=1, column=2, value="InterviewUID")
        sheet.cell(row=1, column=3, value="Status")
        sheet.cell(row=1, column=4, value="Total:")
        sheet.cell(row=1, column=5, value=number_rows)

        for index, row in enumerate(df.itertuples()):
            sheet.append([index + 1, row.InterviewUID, row.Status])

        # подгонка ширины столбцов
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)  # добавляем небольшой запас
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    sheet_title_to_remove = "Sheet"
    sheet_to_remove = new_wb[sheet_title_to_remove]
    new_wb.remove(sheet_to_remove)

    new_wb.save(f"{project_path}\\ID.xlsx")
    new_wb.close()


@error_checker
def get_opens(project_path, project_number, opens_syntax_name, opens_syntax_path):
    if opens_syntax_path is None:
        opens_path = project_path.split("Programming")[0] + "DP\\Opens\\To_coding"
    else:
        opens_path = "\\".join(opens_syntax_path.split("\\")[0:-1])

    print(opens_path)

    if not os.path.exists(opens_path + f"\\{opens_syntax_name}"):
        return 1
        # file_list = os.listdir(opens_path)
        # for file in file_list:
        #     if "opens" and ".sps" in file:
        #         open_syntax_file_list.append(file)

    find_text = r"'...\4. DATA (coding, de, dp)\DP\Opens\To_coding\openq1_to_coding_21-000000-01.xlsx'"
    find_text2 = r'"...\4. DATA (coding, de, dp)\DP\Opens\To_coding\openq1_to_coding_21-000000-01.xlsm"'

    find_text_strip = find_text.strip("'")
    find_text2_strip = find_text2.strip('"')

    pattern = r"\d\d-\d\d\d\d\d\d-\d\d"
    result1 = re.sub(pattern, project_number, find_text_strip)
    result2 = re.sub(pattern, project_number, find_text2_strip)

    result1 = result1.replace(r"...\4. DATA (coding, de, dp)\DP\Opens\To_coding", opens_path)
    result2 = result2.replace(r"...\4. DATA (coding, de, dp)\DP\Opens\To_coding", opens_path)

    if "'" in project_path:
        result1 = '"' + result1 + '"'
        result2 = '"' + result2 + '"'
    else:
        result1 = "'" + result1 + "'"
        result2 = "'" + result2 + "'"

    opens_syntax = open(f"{opens_path}\\{opens_syntax_name}", "r").read()
    opens_syntax = opens_syntax.replace('get file="...".', f'get file="{project_path}\\{project_number}.sav".')
    opens_syntax = opens_syntax.replace("/keep ID", "/keep InterviewID")
    opens_syntax = opens_syntax.replace(find_text, result1)
    opens_syntax = opens_syntax.replace(find_text2, result2)

    new_opens_syntax = open(f"{opens_path}\\{opens_syntax_name}", "w")
    new_opens_syntax.write(opens_syntax)
    new_opens_syntax.close()

    # unicode = "false" charset="UTF-8" в остальных случаях открытые выгружаются некорректно,
    # либо с битой вкладкой статистики, либо с проблемы с кодировкой
    spj_file = open(f"{project_path}\\opens_task.spj", "w")
    syntax_spj = (f'<?xml version="1.0" encoding="UTF-8"?><job codepageSyntaxFiles="false" print="true" '
                  f'syntaxErrorHandling="continue" syntaxFormat="interactive" unicode="false" '
                  f'xmlns="http://www.ibm.com/software/analytics/spss/xml/production" '
                  f'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                  f'xsi:schemaLocation="http://www.ibm.com/software/analytics/spss/xml/production '
                  f'http://www.ibm.com/software/analytics/spss/xml/production/production-1.4.xsd"><locale '
                  f'charset="UTF-8" country="RU" language="ru"/><output outputFormat="viewer" '
                  f'outputPath="test.spv"/><syntax syntaxPath="{opens_path}\\{opens_syntax_name}"/></job>')
    spj_file.write(syntax_spj)
    spj_file.close()

    opens_run_bat = open(f"{project_path}\\opens_run.bat", "w")
    opens_run_bat.write(
        f'"C:\\Program Files\\IBM\\SPSS\\Statistics\\24\\stats.exe" "{project_path}\\opens_task.spj" -production '
        f'silent')
    opens_run_bat.close()

    subprocess.run(f'{project_path}\\opens_run.bat', shell=True)

    opens_file_path = result2.strip("'").strip('"')

    os.remove(f"{project_path}\\opens_run.bat")
    os.remove(f"{project_path}\\opens_task.spj")

    return opens_file_path
